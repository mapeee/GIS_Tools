# -*- coding: cp1252 -*-
#!/usr/bin/python
#-------------------------------------------------------------------------------
# Name:        GIS_Excel
# Purpose:     Exchange of Data betweeen HDF5 and Excel
# Author:      mape
# Created:     25/03/2022
# Copyright:   (c) mape 2022
# Licence:     CC BY-NC 4.0
#-------------------------------------------------------------------------------

import arcpy
import h5py
import numpy as np
import openpyxl as opxl
from openpyxl.styles import Font
import pandas as pd

#--Parameter--#
Database = arcpy.GetParameterAsText(0)
Group = arcpy.GetParameterAsText(1)
dsetHDF5 = arcpy.GetParameterAsText(2).split(";")
Path = arcpy.GetParameterAsText(3)

file5 = h5py.File(Database,'r+')
group5 = file5[Group]
for i in dsetHDF5:
    arcpy.AddMessage("> "+i+"\n")
    data = group5[i]
    a = np.array(data)
    b = pd.DataFrame(a)
    if len(b)<100000:
        path_xls = Path+"\\"+i+".xlsx"
        b.to_excel(path_xls)
        wb = opxl.load_workbook(path_xls)
        sheet = wb.active
        try:
            sheet.cell(1,1,unicode(data.attrs.values()[0], errors='ignore'))
            sheet.cell(1,1).font = Font(color = '00FF0000')
        except: pass #if no attributes defined
        wb.save(path_xls)
        
    else:
        path_xls = Path+"\\"+i+".csv"
        b.to_csv(path_xls)  
        
#--End--#
file5.close()