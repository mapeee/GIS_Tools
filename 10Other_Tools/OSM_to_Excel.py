# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        importing lines for new routes
# Purpose:     Adding Tags to a given set of osm_ids.
# Author:      mape
# Created:     26/07/2023
# Copyright:   (c) mape 2023
# Licence:     	gpl-3.0
#-------------------------------------------------------------------------------

import overpy
import openpyxl
from openpyxl.styles import Font

from pathlib import Path
path = Path.home() / 'python32' / 'python_dir.txt'
path = open(path, mode='r').readlines()
path = path[0] +"\\"+'GIS_Tools'+"\\"+'OSM_to_Excel.txt'
f = open(path, mode='r').read().splitlines()

XLSX_path = f[0]

def Excel(path):
    global wb
    
    XLSX = path
    wb = openpyxl.load_workbook(XLSX)
    sheet= wb.active
    return sheet

def List_osm_id(XLSX):
    List = []
    n = XLSX.max_row -1
    for i in range(n):
        osm_id = XLSX.cell(i+2,1).value
        List.append(osm_id)
    return List

def result_api(osm_IDs):
    osm_IDs = [str(x) for x in osm_IDs]
    string_osm_IDs = ', '.join(osm_IDs)
    query_api = "way(id:"+string_osm_IDs+"); (._;>;); out body;"
    result = overpy.Overpass().query(query_api)
    return result

def write_XLSX(XLSX, osm_IDs, fromrow, results):
    global osmID
    
    tags = ["highway", "surface"]
    if fromrow==0:
        XLSX.cell(1,1).font = Font(bold = True)
        for i, tag in enumerate(tags):
            XLSX.cell(1,i+2).value = tag
            XLSX.cell(1,i+2).font = Font(bold = True)
    
    for i, osmID in enumerate(osm_IDs):
        for e, tag in enumerate(tags):
            try: XLSX.cell(i+2+fromrow,e+2).value = results.get_way(osmID).tags.get(tag, "n/a")
            except: XLSX.cell(i+2+fromrow,e+2).value = "osm_id missing"

#--Parameter--#
XLSX = Excel(XLSX_path)
osm_IDs = List_osm_id(XLSX)
for i in range(int(len(osm_IDs)/5000)+1):
    row = [i*5000, (i+1)*5000]
    osm_data = result_api(osm_IDs[row[0]:row[1]])
    write_XLSX(XLSX, osm_IDs[row[0]:row[1]], row[0], osm_data)

wb.save(XLSX_path)


