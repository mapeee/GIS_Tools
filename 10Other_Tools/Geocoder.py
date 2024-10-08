#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      mape
#
# Created:     23/11/2017
# Copyright:   (c) mape 2017
# Licence:     <your licence>
#-------------------------------------------------------------------------------
from opencage.geocoder import OpenCageGeocode
import geocoder
import openpyxl
import requests
from pyproj import Transformer
import time
start_time = time.time()
from pathlib import Path
f = open(Path.home() / 'python32' / 'python_dir.txt', mode='r')
for i in f: path = i
path = Path.joinpath(Path(path),'GIS_Tools','Geocoder.txt')
f = path.read_text().split('\n')

#--connecting to opencagegeocode--#
osm_key = f[1]
bing_key = f[2]
here_key = f[3]
geocoder_osm = OpenCageGeocode(osm_key)

#--values--#
Street = 3
ZIP = 4
City = 5

#--writer--#
x,y,coder = 12, 13, 14

#--open excel--#
excelfile = f[0]
wbk = openpyxl.load_workbook(excelfile)
wks = wbk.worksheets[0]

def bing(query):
    try: 
        result_bing = geocoder.bing(query, key = bing_key)
        if result_bing.postal == None or result_bing.postal not in query or result_bing.raw["confidence"] != "High":
            return [None]
        else: return result_bing.lng, result_bing.lat, 'bing'
    except: return [None]

def geocage(query):
    result_osm = geocoder_osm.geocode(query)
    for i in result_osm:
        try:
            lat = i["geometry"]["lat"]
            lng = i["geometry"]["lng"]
            break
        except: return [None]
    return lng, lat, 'opencage'

def here(query):
    URL = "https://geocode.search.hereapi.com/v1/geocode"
    PARAMS = {'apikey':here_key,'q':query}
    try:
        r = requests.get(url = URL, params = PARAMS)
        data = r.json()
        lat = data['items'][0]['position']['lat']
        lng = data['items'][0]['position']['lng']
        return lng, lat, 'here'
    except:
        return 0, 0, 'error'

def transform(XY):
    epsg_in = "epsg:4326"
    epsg_out = "epsg:25832"
    transformer = Transformer.from_crs(epsg_in, epsg_out, always_xy=True)
    x_out, y_out = transformer.transform(XY[0], XY[1])
    return x_out, y_out

def exl_writer(results, row):
    wks.cell(row, x).value = transform(results)[0]
    wks.cell(row, y).value = transform(results)[1]
    wks.cell(row, coder).value = results[2]


#--geocoding--#
for row in range(2,wks.max_row+1):##to prevent 0 and no header
    if wks.cell(row,Street).value == None: break

    query = wks.cell(row,Street).value+", "+str(wks.cell(row,ZIP).value)+" "+wks.cell(row,City).value
    #bing
    result = bing(query)
    #geocage
    if result[0] == None: result = geocage(query)
    #here
    if result[0] == None: result = here(query)

    exl_writer(result,row)
    time.sleep(0.1) ##pause 1.5 seconds
    if row in list(range(0,wks.max_row+1,10)): print(row)

#--header--#
wks.cell(1, x).value = 'X'
wks.cell(1, y).value = 'Y'
wks.cell(1, coder).value = 'geocoder'

#--end--#
wbk.save(excelfile)
wbk.close()
seconds = int(time.time() - start_time)
print("--finished after ",seconds,"seconds--")